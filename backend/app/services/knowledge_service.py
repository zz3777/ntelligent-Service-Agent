import uuid
from io import BytesIO

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from app.models.knowledge_base import Document, DocumentChunk


async def extract_text_from_file(content: bytes, file_type: str) -> str:
    if file_type == "pdf":
        return _extract_pdf_text(content)
    elif file_type == "docx":
        return _extract_docx_text(content)
    elif file_type == "xlsx":
        return _extract_xlsx_text(content)
    elif file_type == "image":
        return _extract_image_text(content)
    return ""


def _extract_pdf_text(content: bytes) -> str:
    # 1. 先用 pdfplumber（能力最强）
    text = _extract_pdf_with_pdfplumber(content)
    if len(text.strip()) > 50:
        return text

    # 2. 再试 PyPDF2
    text = _extract_pdf_with_pypdf2(content)
    if len(text.strip()) > 50:
        return text

    # 3. 都不行，说明是扫描件，用 pdf2image + pytesseract 做 OCR
    return _extract_pdf_with_ocr(content)


def _extract_pdf_with_pdfplumber(content: bytes) -> str:
    try:
        import pdfplumber
        texts = []
        with pdfplumber.open(BytesIO(content)) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    texts.append(text)
        return "\n".join(texts)
    except Exception:
        return ""


def _extract_pdf_with_pypdf2(content: bytes) -> str:
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(BytesIO(content))
        return "\n".join(page.extract_text() or "" for page in reader.pages)
    except Exception:
        return ""


def _extract_pdf_with_ocr(content: bytes) -> str:
    """对扫描件 PDF 逐页转图片后 OCR"""
    try:
        from pdf2image import convert_from_bytes
        import pytesseract
        import tempfile, os

        images = convert_from_bytes(content, dpi=200)
        all_texts = []
        for img in images:
            text = pytesseract.image_to_string(img, lang="chi_sim+eng")
            if text.strip():
                all_texts.append(text.strip())

        return "\n".join(all_texts) if all_texts else "[OCR 未能识别文字]"
    except Exception as e:
        return f"[PDF OCR 提取失败: {e}]"


def _extract_docx_text(content: bytes) -> str:
    try:
        from docx import Document as DocxDocument
        doc = DocxDocument(BytesIO(content))
        return "\n".join(p.text for p in doc.paragraphs)
    except Exception as e:
        return f"[DOCX 提取失败: {e}]"


def _extract_xlsx_text(content: bytes) -> str:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(content), read_only=True)
        texts = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                texts.append(" | ".join(str(c) for c in row if c is not None))
        return "\n".join(texts)
    except Exception as e:
        return f"[XLSX 提取失败: {e}]"


def _extract_image_text(content: bytes) -> str:
    try:
        from paddleocr import PaddleOCR
        ocr = PaddleOCR(use_angle_cls=True, lang="ch", show_log=False)
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as f:
            f.write(content)
            tmp_path = f.name
        result = ocr.ocr(tmp_path)
        texts = []
        for line_group in result or []:
            for line in line_group or []:
                texts.append(line[1][0])
        import os
        os.unlink(tmp_path)
        return "\n".join(texts)
    except Exception as e:
        return f"[OCR 提取失败: {e}]"


def _chunk_text(text: str, chunk_size: int = 300, overlap: int = 50) -> list[str]:
    if not text.strip():
        return []

    import re

    # 1. 先按章节标题分割（中文简历常见标题）
    section_patterns = [
        r"\n(?=教育经历)", r"\n(?=专业技能)", r"\n(?=项目经历)",
        r"\n(?=工作经历)", r"\n(?=自我评价)", r"\n(?=个人信息)",
        r"\n(?=实习经历)", r"\n(?=获奖情况)", r"\n(?=证书资质)",
        r"\n(?=基本信息)", r"\n(?=求职意向)", r"\n(?=语言能力)",
    ]
    sections = text
    for pattern in section_patterns:
        sections = re.sub(pattern, "\n===SPLIT===\n", sections)

    if "===SPLIT===" in sections:
        parts = [p.strip() for p in sections.split("===SPLIT===") if p.strip()]
    else:
        # 2. 没有章节标题，按段落和换行分割
        parts = re.split(r"\n\s*\n|\n", text)
        parts = [p.strip() for p in parts if p.strip()]

    # 3. 对过长的段落再做子分片
    chunks = []
    for part in parts:
        if len(part) <= chunk_size:
            chunks.append(part)
        else:
            # 按句子边界分
            sentences = re.split(r"(?<=[。！？；\n])", part)
            current = ""
            for sent in sentences:
                if len(current) + len(sent) < chunk_size:
                    current += sent
                else:
                    if current:
                        chunks.append(current.strip())
                    current = sent
            if current:
                chunks.append(current.strip())

    return chunks or [text.strip()]


async def process_document(doc_id: uuid.UUID, text: str, db: AsyncSession) -> int:
    from app.agent.rag import add_to_knowledgebase, get_chroma_client
    from sqlalchemy import delete

    # 先删旧分片
    await db.execute(delete(DocumentChunk).where(DocumentChunk.document_id == doc_id))
    try:
        client = get_chroma_client()
        collection = client.get_or_create_collection(name="enterprise_knowledge")
        old = collection.get(where={"document_id": str(doc_id)})
        if old and old["ids"]:
            collection.delete(ids=old["ids"])
    except Exception:
        pass

    chunks = _chunk_text(text)
    doc_id_str = str(doc_id)

    for i, chunk_content in enumerate(chunks):
        chunk = DocumentChunk(
            document_id=doc_id,
            content=chunk_content,
            chunk_index=i,
        )
        db.add(chunk)

        try:
            await add_to_knowledgebase(chunk_content, doc_id_str, i)
        except Exception:
            pass

    await db.commit()
    return len(chunks)
